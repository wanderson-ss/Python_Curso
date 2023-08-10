import openpyxl
from openpyxl import Workbook
from tkinter import *
from tkinter import filedialog
import requests
import cv2
import numpy as np
import imutils
import pytesseract
from PIL import Image


def criar():
    nome = entrada.get()

    pasta = openpyxl.load_workbook(nome + '.xlsx')
    planilha = pasta.active

    img = filedialog.askopenfilename(title='Selecione o arquivo:',
                                     filetypes=[('Arquivos de imagens', '*.png;*.jpg;*.jpeg')])

    imagem = Image.open(img)
    texto = pytesseract.image_to_string(imagem, lang='por')

    linhas = texto.split('/n')
    for x, linha in enumerate(linhas):
        colunas = linha.split()
    for y, coluna in enumerate(colunas):
        celula = planilha.cell(row=x + 1, column=y + 1)
        celula.value = coluna

    img = entrada.get()
    pasta.save(nome + '.xlsx')


def leitura():
    url = "http://192.168.15.7:8080/shot.jpg"

    while True:
        img_resp = requests.get(url)
        img_arr = np.array(bytearray(img_resp.content), dtype=np.uint8)
        img = cv2.imdecode(img_arr, -1)
        img = imutils.resize(img, width=1000, height=1800)
        if cv2.waitKey(1) == ord('s'):
            cv2.imwrite('image.png', img)
        cv2.imshow("Android_cam", img)

        if cv2.waitKey(1) == 27:
            break
    cv2.destroyAllWindows()

    # Defina o caminho para o executável do Tesseract
    pytesseract.pytesseract.tesseract_cmd = r'C:\Program Files\Tesseract-OCR\tesseract.exe'

    # Converta a imagem para escala de cinza
    gray = cv2.cvtColor(img, cv2.COLOR_BGR2GRAY)
    blur = cv2.blur(gray, (5, 5))

    # Aplique limiarização para remover o ruído de fundo
    _, thresh = cv2.threshold(blur, 0, 255, cv2.THRESH_BINARY_INV + cv2.THRESH_OTSU)

    # Aplique dilatação e erosão para conectar os caracteres quebrados e separar os caracteres colados
    kernel = cv2.getStructuringElement(cv2.MORPH_RECT, (3, 3))
    thresh = cv2.morphologyEx(thresh, cv2.MORPH_CLOSE, kernel)

    # Extraia o texto da imagem pré-processada
    text = pytesseract.image_to_string(thresh, config='--psm 6')
    print(text)
    print()

    # Use expressões regulares para encontrar os atributos e seus valores no texto
    import re
    attributes = re.findall(r'(\w+): (.+)', text)
    print(attributes)

    # Crie um novo arquivo do Excel
    wb = Workbook()
    ws = wb.create_sheet("Dados")

    ws.append(["Nome", "Nome do Pai", "Nome da Mãe", "Endereço", "Telefone", "Batismo", "Sexo"])

    tupla = dict(attributes)

    row = [tupla.get("Nome", ""), tupla.get("Nome do Pai", ""), tupla.get("Nome da Mãe", ""), tupla.get("Endereço", ""),
           tupla.get("Telefone", ""), tupla.get("Batismo", ""), tupla.get("Sexo", "")]
    ws.append(row)

    wb.save("dados.xlsx")


janela = Tk()

janela.title("Bem-Vindo")
janela.geometry("800x300")

titulo = Label(janela, text="Escolha alguma opção")
titulo.pack()

rotulo = Label(janela, text='Escolha  um nome para o arquivo:')
rotulo.pack()

entrada = Entry(janela)
entrada.pack()

botao = Button(janela, text="criar manualmente", command=criar)
botao.pack()

botao1 = Button(janela, text="criar usando imagens", command=leitura)
botao1.pack()

janela.mainloop()